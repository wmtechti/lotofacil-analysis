"""
Exportador para Excel - Análise de Loterias

Exporta todos os dados e análises para formato Excel (.xlsx) com múltiplas abas,
facilitando a visualização e uso das análises.
"""

import pandas as pd
from pathlib import Path
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Exportador de análises para Excel"""
    
    def __init__(self):
        self.output_file = 'out/LOTOFACIL_ANALISE_COMPLETA.xlsx'
        
    def export_lotofacil_games(self, writer):
        """Exportar jogos otimizados da Lotofácil"""
        print("📄 Exportando jogos otimizados (30 jogos)...")
        
        df = pd.read_csv('out/jogos_otimizados_combined.csv')
        df = df[['jogo_id', 'estrategia', 'numeros_str']]
        df.columns = ['ID', 'Estratégia', 'Números']
        
        df.to_excel(writer, sheet_name='Jogos Otimizados (30)', index=False)
        
    def export_lotofacil_games_100(self, writer):
        """Exportar jogos adicionais (100 jogos)"""
        print("📄 Exportando jogos adicionais (100 jogos)...")
        
        df = pd.read_csv('out/jogos_otimizados_100.csv')
        df = df[['jogo_id', 'estrategia', 'numeros_str']]
        df.columns = ['ID', 'Estratégia', 'Números']
        
        df.to_excel(writer, sheet_name='Jogos Otimizados (100)', index=False)
        
    def export_backtesting_results(self, writer):
        """Exportar resultados do backtesting"""
        print("📄 Exportando resultados do backtesting...")
        
        # Resultados por jogo
        df_games = pd.read_csv('out/backtesting/resultados_por_jogo.csv')
        df_games = df_games[['jogo_id', 'strategy', 'avg_matches', 'prize_rate', 
                             'prizes_11', 'prizes_12', 'prizes_13', 'prizes_14', 'prizes_15', 'numbers']]
        df_games.columns = ['ID', 'Estratégia', 'Média Acertos', 'Taxa Prêmio %', 
                           '11 Acertos', '12 Acertos', '13 Acertos', '14 Acertos', '15 Acertos', 'Números']
        df_games.to_excel(writer, sheet_name='Backtesting - Jogos', index=False)
        
        # Resultados por estratégia
        df_strat = pd.read_csv('out/backtesting/resultados_por_estrategia.csv')
        df_strat.columns = ['Estratégia', 'Qtd Jogos', 'Média Acertos', 'Melhor Acerto',
                           'Taxa Prêmio %', '11 Acertos', '12 Acertos', '13 Acertos', '14 Acertos', '15 Acertos']
        df_strat.to_excel(writer, sheet_name='Backtesting - Estratégias', index=False)
        
    def export_lotofacil_analysis(self, writer):
        """Exportar análises da Lotofácil"""
        print("📄 Exportando análises da Lotofácil...")
        
        # Números quentes/frios
        df_hot = pd.read_csv('out/lotofacil/numeros_quentes_frios.csv')
        df_hot.columns = ['Número', 'Frequência', 'Esperado', 'Desvio %', 'Categoria']
        df_hot.to_excel(writer, sheet_name='Números Quentes-Frios', index=False)
        
        # Super pares
        df_pairs = pd.read_csv('out/lotofacil/pares_forca.csv')
        df_pairs = df_pairs[df_pairs['categoria'].str.contains('Super Par|Forte', na=False)]
        df_pairs.columns = ['Número A', 'Número B', 'Aparições', 'Força %', 'Categoria']
        df_pairs.to_excel(writer, sheet_name='Super Pares', index=False)
        
        # Frequência por linhas
        df_lines = pd.read_csv('out/lotofacil/freq_linhas.csv')
        df_lines.columns = ['Linha', 'Frequência']
        df_lines.to_excel(writer, sheet_name='Frequência Linhas', index=False)
        
        # Frequência por colunas
        df_cols = pd.read_csv('out/lotofacil/freq_colunas.csv')
        df_cols.columns = ['Coluna', 'Frequência']
        df_cols.to_excel(writer, sheet_name='Frequência Colunas', index=False)
        
    def export_megasena_summary(self, writer):
        """Exportar resumo da análise Mega-Sena"""
        print("📄 Exportando resumo Mega-Sena...")
        
        # Carregar JSON
        with open('out/megasena/megasena_analyses.json', 'r', encoding='utf-8') as f:
            analyses = json.load(f)
        
        # Criar DataFrame resumido
        summary_data = []
        for a in analyses[:100]:  # Primeiros 100 para não ficar muito grande
            # Calcular distribuição por quadrante (percentuais)
            quadrants = a['regions']['quadrants']
            total_nums = sum(quadrants.values())
            
            summary_data.append({
                'Concurso': a['concurso'],
                'Números': str(sorted(a['numeros'])),
                'Padrão': a['pattern'],
                'Pares Adjacentes': a['contiguity']['connected_pairs'],
                'Dispersão Média': round(a['dispersion']['mean_pairwise_distance'], 2),
                'Q1': quadrants['Q1'],
                'Q2': quadrants['Q2'],
                'Q3': quadrants['Q3'],
                'Q4': quadrants['Q4']
            })
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Mega-Sena (Amostra)', index=False)
        
    def export_combined_insights(self, writer):
        """Exportar insights da análise combinada"""
        print("📄 Exportando insights combinados...")
        
        # Criar DataFrame com insights principais
        insights_data = [
            {
                'Categoria': 'Dispersão Espacial',
                'Mega-Sena': '78.13% DISPERSO | Dist. média: 5.34',
                'Lotofácil': 'Dist. média: 2.41',
                'Estratégia': 'Distribuir números por DIFERENTES regiões do grid',
                'Score': '8.5/10'
            },
            {
                'Categoria': 'Equilíbrio Regional',
                'Mega-Sena': '~25% por quadrante (PERFEITO)',
                'Lotofácil': 'Balanceado em linhas/colunas',
                'Estratégia': 'Selecionar números de TODAS as regiões',
                'Score': '9.0/10'
            },
            {
                'Categoria': 'Contiguidade',
                'Mega-Sena': '0.87 pares adjacentes | 38.69% dispersos',
                'Lotofácil': 'Baixa adjacência confirmada',
                'Estratégia': 'EVITAR muitos números vizinhos (max 1-2 pares)',
                'Score': '7.5/10'
            },
            {
                'Categoria': 'Co-ocorrência',
                'Mega-Sena': 'N/A',
                'Lotofácil': '43 super pares | Par top: 11-20 (1362x)',
                'Estratégia': 'INCLUIR 1-2 super pares nos jogos (Lotofácil)',
                'Score': '6.5/10'
            },
            {
                'Categoria': 'Aleatoriedade',
                'Mega-Sena': 'Confirmada (alta dispersão)',
                'Lotofácil': 'Desvios <5% (confirmada)',
                'Estratégia': 'NÃO confiar em números "quentes" de curto prazo',
                'Score': '3.0/10'
            }
        ]
        
        df_insights = pd.DataFrame(insights_data)
        df_insights.to_excel(writer, sheet_name='Insights Combinados', index=False)
        
    def export_recommendations(self, writer):
        """Exportar recomendações finais"""
        print("📄 Exportando recomendações...")
        
        recommendations_data = [
            {'Tipo': 'FAZER ✅', 'Recomendação': 'Distribuir números por TODAS as regiões do grid', 'Prioridade': 'ALTA'},
            {'Tipo': 'FAZER ✅', 'Recomendação': 'Incluir 1-2 super pares (Lotofácil)', 'Prioridade': 'ALTA'},
            {'Tipo': 'FAZER ✅', 'Recomendação': 'Balancear ímpares/pares (7-8 cada)', 'Prioridade': 'MÉDIA'},
            {'Tipo': 'FAZER ✅', 'Recomendação': 'Garantir dispersão espacial', 'Prioridade': 'ALTA'},
            {'Tipo': 'FAZER ✅', 'Recomendação': 'Evitar concentração em uma única área', 'Prioridade': 'ALTA'},
            {'Tipo': 'EVITAR ❌', 'Recomendação': 'Muitos números adjacentes (vizinhos)', 'Prioridade': 'ALTA'},
            {'Tipo': 'EVITAR ❌', 'Recomendação': 'Sequências óbvias (1,2,3,4,5...)', 'Prioridade': 'MÉDIA'},
            {'Tipo': 'EVITAR ❌', 'Recomendação': 'Concentração em bordas ou centro', 'Prioridade': 'MÉDIA'},
            {'Tipo': 'EVITAR ❌', 'Recomendação': 'Confiar apenas em números "quentes"', 'Prioridade': 'ALTA'},
            {'Tipo': 'EVITAR ❌', 'Recomendação': 'Padrões visuais óbvios (diagonais, cruzes)', 'Prioridade': 'MÉDIA'}
        ]
        
        df_rec = pd.DataFrame(recommendations_data)
        df_rec.to_excel(writer, sheet_name='Recomendações', index=False)
        
        # Adicionar resumo de expectativas
        expectativas_data = [
            {'Métrica': 'Taxa de prêmio (estratégias otimizadas)', 'Valor': '~12.4%', 'Observação': 'Lotofácil'},
            {'Métrica': 'Taxa de prêmio (baseline aleatório)', 'Valor': '~11.3%', 'Observação': 'Lotofácil'},
            {'Métrica': 'Ganho potencial', 'Valor': '+1.08%', 'Observação': 'Modesto mas consistente'},
            {'Métrica': 'Melhor estratégia', 'Valor': 'Equilíbrio Regional', 'Observação': 'Score 9.0/10'},
            {'Métrica': 'Segunda melhor', 'Valor': 'Dispersão Máxima', 'Observação': 'Score 8.5/10'}
        ]
        
        df_exp = pd.DataFrame(expectativas_data)
        df_exp.to_excel(writer, sheet_name='Expectativas', index=False)
        
    def apply_formatting(self):
        """Aplicar formatação ao arquivo Excel"""
        print("🎨 Aplicando formatação...")
        
        wb = load_workbook(self.output_file)
        
        # Cores
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Formatar cabeçalho
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Congelar primeira linha
            ws.freeze_panes = 'A2'
        
        wb.save(self.output_file)
        
    def export_all(self):
        """Exportar tudo para Excel"""
        print("=" * 80)
        print("📊 EXPORTANDO ANÁLISES PARA EXCEL")
        print("=" * 80)
        
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            # Exportar cada aba
            self.export_lotofacil_games(writer)
            self.export_lotofacil_games_100(writer)
            self.export_backtesting_results(writer)
            self.export_lotofacil_analysis(writer)
            self.export_megasena_summary(writer)
            self.export_combined_insights(writer)
            self.export_recommendations(writer)
        
        # Aplicar formatação
        self.apply_formatting()
        
        print("\n" + "=" * 80)
        print(f"✅ EXPORTAÇÃO CONCLUÍDA!")
        print(f"📁 Arquivo salvo: {self.output_file}")
        print("=" * 80)
        
        # Listar abas criadas
        wb = load_workbook(self.output_file)
        print(f"\n📋 Abas criadas ({len(wb.sheetnames)}):")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"   {i}. {sheet_name}")


def main():
    """Executar exportação"""
    exporter = ExcelExporter()
    exporter.export_all()


if __name__ == '__main__':
    main()
